from pathlib import Path
from openpyxl import load_workbook
from mysite import settings
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from bs4 import BeautifulSoup
# from test_email import send_email
# from get_shopify_theme import get_shopify_details
import urllib.request
import urllib.parse
import random
import json
import os
import requests
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
import smtplib

def get_psi_scores(url: str, strategies: list=["mobile", "desktop"]) -> list:
    scores = []
    api_keys = ["AIzaSyCiGfKHVbQiOUi2pcBkE8Jl276JtAIVLVA", "AIzaSyDIYtnQE0MTmlDqgRzy1yGopRAlxrWM75M", "AIzaSyBnWTnurdDnirNgycqRctHJzha-RwHs_6Q", "AIzaSyA6jl7rZlyarDVLE9IiSjybezQYIdXlkag", "AIzaSyDPSoydseAmWfCi_0-ZZrleqKXOz2VJi8M", "AIzaSyARQlzDXTxGd7hquGmGnb9GZN02tOBvAaQ"]
    api_key = random.choice(api_keys)
    # print("API Key used ------------->", api_key)
    for strategy in strategies:
        # try:
        # print(f'https://www.googleapis.com/pagespeedonline/v5/runPagespeed?url={url}/&strategy={strategy}&key={api_key}')
        data = urllib.request.urlopen(f'https://www.googleapis.com/pagespeedonline/v5/runPagespeed?url={url}/&strategy={strategy}&key={api_key}').read().decode('UTF-8')
        # data = requests.get(f'https://www.googleapis.com/pagespeedonline/v5/runPagespeed?url={url}/&strategy={strategy}&key={api_key}')#.read().decode('UTF-8')
        performance = int(float(json.loads(data)["lighthouseResult"]["categories"]["performance"]["score"])*100)
        scores.append(str(performance))
        if strategy == "mobile":
            raw_lcp_score = json.loads(data)["lighthouseResult"]["audits"]["largest-contentful-paint"]["displayValue"]
            temp_var = raw_lcp_score.split(".")
            lcp_score = f"{temp_var[0]}.{temp_var[1][0]}"
            scores.append(lcp_score)
        # except HTTPError:
        #     print("Too many requests! Please try again in some time.")
    return scores

def get_shopify_boomerang_details(shopify_url: str, key: str):
    """
    :param shopify_url: URL to the shopify store.
    :param key: Specific detail to extract.\n
    Keys can be following:\n
    Name -> Theme name.\n
    Version -> Theme version.
    """
    url = shopify_url
    response = requests.get(url)
    html_content = response.content
    soup = BeautifulSoup(html_content, 'html.parser')
    visible_scripts = soup.find_all('script', attrs={'class':'boomerang', 'src': False})
    for script in visible_scripts:
        js_code = script.string
        if js_code is not None and "window.BOOMR" in js_code:
            # print(js_code.strip().split("\n"))
            for i in js_code.strip().split("\n"):
                if f"window.BOOMR.theme{key.capitalize()}" in i.lstrip():
                # if re.search(f"window.BOOMR.{key}", i): # theme{key.capitalize()
                    # print("found")
                    return i.lstrip().split("=")[1].lstrip().lstrip('"').rstrip('";')

def send_email(sender_email: str, sender_password: str, to: list, mobile_score, lcp_score, desktop_score) -> None:
    mail_id: list = to
    email = sender_email # ""
    password = sender_password # ""
    message = MIMEMultipart('alternative')
    mail= smtplib.SMTP_SSL("gsgp1032.siteground.asia", 465)
    mail.login(email, password)
    message['From'] = email
    message['To'] = mail_id[0]
    message['Subject'] = "Test"
    store_dtop_psi = desktop_score
    store_mob_psi = mobile_score
    lcp_score = lcp_score
    html = f"""
    <html>
        <body>
            <p>Hello there!</p>
            <p>I hope this message finds you well. I'm reaching out because I'm impressed with your brand and I believe I can help you improve your online store’s speed which usually results in higher conversion.</p>
            <p>We have conducted a quick analysis of your website's speed on mobile and desktop and noticed that there is room for improvement. On desktop, as per the google page speed index, your website's speed is {store_dtop_psi}, and on mobile, it's {store_mob_psi} and your LCP score is {lcp_score}.</p>
            <p>As per industry standards, we recommend a speed score of at least 80 on desktop and 40 on mobile devices.</p>
            <p>At <a href=https://www.propero.in/>Propero</a>, we specialize in website speed optimization and have helped many D2C brands like yours such as OyeHappy, Bombay Sweet Shop, Stuffcool, and many more to improve website performance.</p>
            <p>By performing a comprehensive website audit, we can identify potential speed optimization issues and provide customized solutions to enhance your website's speed and performance.</p>
            <p>Some of the benefits of improved speed are:<br>
            &emsp;●&ensp;Low bounce rate resulting in higher conversion.<br>
            &emsp;●&ensp;Higher search engine rankings and online visibility.<br>
            &emsp;●&ensp;Enhanced user experience and satisfaction.<br>
            </p>
            <p>We'd love to discuss your website speed optimization needs further and show you how we can help you achieve your goals. Please let us know your availability for a quick call.</p>
            <p>Thank you for your time, and we look forward to hearing from you soon.</p>
            <p>Best regards,<br>Garima</p>
            <p>PS: you can check speed scores from multiple sources by using our free analyzer tool <a href=https://speed.propero.in>https://speed.propero.in</a></p>
        <body>
    <html>
    """
    message.attach(MIMEText(html, 'html'))
    text = message.as_string()
    mail.sendmail(email, mail_id, text)
    mail.quit
    print(f"Mail sent to {mail_id} Sucessfully")

def send_email_to_myself(sender_email, sender_password, to, file_path):
    mail_id: list = to
    email = sender_email
    password = sender_password
    message = MIMEMultipart()
    mail = smtplib.SMTP_SSL("gsgp1032.siteground.asia", 465)
    mail.login(email, password)
    message['From'] = email
    message['To'] = mail_id[0]
    message['Subject'] = "Test"

    # Attach Excel file
    with open(file_path, "rb") as f:
        attach = MIMEApplication(f.read(),_subtype="xlsx")
        attach.add_header('Content-Disposition','attachment',filename=str(file_path))
        message.attach(attach)

    text = message.as_string()
    mail.sendmail(email, mail_id, text)
    mail.quit
    print(f"Mail sent to {mail_id} Sucessfully")


def process_file(input_file, output_file, name):
    if not os.path.exists(settings.MEDIA_ROOT):
        os.mkdir(settings.MEDIA_ROOT)
    if not os.path.exists(settings.BASE_DIR / "storage"):
        os.mkdir(settings.BASE_DIR / "storage")
    wb = load_workbook(input_file)
    sh = wb.active
    sh: Worksheet
    m_row = sh.max_row
    sh.insert_cols(idx=4)
    sh.insert_cols(idx=5)
    sh.insert_cols(idx=6)
    sh.cell(1, 4).value = "Mobile score"
    sh.cell(1, 5).value = "LCP score"
    sh.cell(1, 6).value = "Desktop score"
    sh.cell(1, 7).value = "Shopify Base Theme"
    sh.cell(1, 8).value = "Shopify Theme Version"
    sh.cell(1, 9).value = "Status"
    sh.cell(1, 10).value = "Exception"
    bold_font = Font(bold=True)
    for num in range(4, 11):
        sh.cell(1, num).font = bold_font
    wb.save(output_file)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet: Worksheet
    for i in range(2, m_row+1):
        cell_obj = sheet.cell(row=i, column=3)
        website = cell_obj.value
        p = urllib.parse.urlparse(website, 'https')
        netloc = p.netloc or p.path
        path = p.path if p.netloc else ''
        if not netloc.startswith('www.'):
            netloc = 'www.' + netloc
        p = urllib.parse.ParseResult('https', netloc, path, *p[3:])
        url = p.geturl()
        print(f"Analyzing {url}...")
        emails = []
        scores = get_psi_scores(url)
        for col in range(4, 7):
            sheet.cell(i, col).value = scores[col-4]
        email_data = sheet.cell(i, 2).value
        if email_data is None:
            sheet.cell(i, 9).value = "Not processed"
            sheet.cell(i, 10).value = "No email found"
            print(f"Did not process for site {url} as no email found.")
        else:
            if ":" in email_data:
                for email in email_data.split(":"):
                    emails.append(email)
            else:
                emails.append(email_data)
            shopify_base_theme = get_shopify_boomerang_details(url, "name")
            shopify_theme_version = get_shopify_boomerang_details(url, "version")
            if shopify_base_theme is None or shopify_base_theme == "":
                shopify_base_theme = "Not available"
            if shopify_theme_version is None or shopify_theme_version == "":
                shopify_theme_version = "Not available"
            print(f"Score for {url} are {scores}, Shopify base theme is {shopify_base_theme} and theme version is {shopify_theme_version}.")
            sheet.cell(i, 7).value = shopify_base_theme
            sheet.cell(i, 8).value = shopify_theme_version
            send_email("speed@propero.in", "Propero$345", ["suyashdhir@gmail.com"], scores[0], scores[1], scores[2])
            sheet.cell(i, 9).value = "Mail sent"
        workbook.save(output_file)
    send_email_to_myself("speed@propero.in", "Propero$345", ["suyashdhir@gmail.com"], settings.BASE_DIR / "storage" / name)
    os.remove(os.path.join(settings.MEDIA_ROOT, name))
    os.remove(settings.BASE_DIR / "storage" / name)